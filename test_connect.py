from netmiko import ConnectHandler
from openpyxl import load_workbook
import sys

load_WB = load_workbook("sample.xlsx") # 엑셀 파일 명 넣기
load_WS1 = load_WB['Machines_IP'] # 엑셀 시트 명 넣기 (IP)
load_WS2 = load_WB['config_command'] # 엑셀 시트 명 넣기 (config)

# config_command = ["do ter len 0", "do show version", "do show arp",
#                   "do show ip int br", "do show int status", "do show ip route"]

for x in range(1, load_WS1.max_row + 1):
    protocol = load_WS1.cell(row=x, column=1).value
    road_IP = load_WS1.cell(row=x, column=2).value
    road_ID = load_WS1.cell(row=x, column=3).value
    road_PW = load_WS1.cell(row=x, column=4).value
    road_enable = load_WS1.cell(row=x, column=5).value

    try:
        cisco_test = {
            'device_type': protocol,
            'ip': road_IP,
            'username': road_ID,
            'password': road_PW,
            'secret': road_enable,
            }
        terminal = ConnectHandler(**cisco_test)
        terminal.enable()

        sys.output_text = open(road_IP + "_log.txt", "w")
        result = terminal.find_prompt()
        print('연결된 장비 : ' + road_IP)
        print('연결된 장비 : ' + road_IP, file=sys.output_text)
        print(result)
        print(result, file=sys.output_text)
        for send_config in range(1, load_WS2.max_row + 1):
            config_set = load_WS2.cell(row = send_config, column=1).value
            # output = terminal.send_config_set(config_set) ## 장비 셋팅
            output = terminal.send_command(config_set) ##장비 설정 확인
            print("\n" + "-"*35 + config_set + "-"*35 + "\n")
            print("\n" + "-"*35 + config_set + "-"*35 + "\n", file=sys.output_text)
            print(output)
            print(output, file=sys.output_text)
        sys.output_text.close()

        terminal.disconnect()

    except Exception as error:
        sys.error_text = open(road_IP + "_error.txt", "w")
        print(error, file=sys.error_text)

    print('연결 종료 : ' + road_IP)

