from operator import itemgetter
from netmiko import ConnectHandler
from openpyxl import load_workbook, Workbook
import sys

load_WB = load_workbook("sample.xlsx") # 엑셀 파일 명 넣기
load_WS1 = load_WB['Machines_IP'] # 엑셀 시트 명 넣기 (IP)
load_WS2 = load_WB['config_command'] # 엑셀 시트 명 넣기 (config)

write_wb = Workbook() # excel 활성화
write_ws1 = write_wb.active # sheet 활성화
write_ws1.title = "Create_Inventory" # sheet name 변경

# config_command = ["do ter len 0", "do show version", "do show arp",
#                   "do show ip int br", "do show int status", "do show ip route"]
write_ws1.append(['번호', 'Hostname', 'IP', 'Version', 'Serial Number'])
for x in range(1, load_WS1.max_row + 1):
    protocol = load_WS1.cell(row=x, column=1).value
    road_IP = load_WS1.cell(row=x, column=2).value
    road_ID = load_WS1.cell(row=x, column=3).value
    road_PW = load_WS1.cell(row=x, column=4).value
    road_enable = load_WS1.cell(row=x, column=5).value
    try:
        cisco_test = {
            'device_type': protocol,
            'ip': road_IP,
            'username': road_ID,
            'password': road_PW,
            'secret': road_enable,
            }
        terminal = ConnectHandler(**cisco_test)
        terminal.enable()

        result = terminal.find_prompt()
        print('연결된 장비 : ' + road_IP)
        print(result)
        output1 = terminal.send_command("show version", use_textfsm=True)
        output2 = terminal.send_command("show inventory", use_textfsm=True)

        show_version = output1[(0)]
        show_inventory = output2[(0)]
        gethostname = itemgetter('hostname')
        wr_hostname = gethostname(show_version)
        getVersion = itemgetter('version')
        wr_Version = getVersion(show_version)
        getserial = itemgetter('sn')
        wr_serial = getserial(show_inventory)
        write_ws1.append([x, wr_hostname, road_IP, wr_Version, wr_serial])
        print(x, wr_hostname, road_IP, wr_Version, wr_serial)
        terminal.disconnect()
        print('연결 종료 : ' + road_IP)
    except Exception as error:
        sys.error_text = open(road_IP + "_error.txt", "w")
        print(error, file=sys.error_text)
    write_wb.save('Create_Inventory.xlsx')
    
