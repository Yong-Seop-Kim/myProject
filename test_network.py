from operator import itemgetter
from netmiko import ConnectHandler
from openpyxl import load_workbook
import sys

load_WB = load_workbook("sample.xlsx") # 엑셀 파일 명 넣기
load_WS1 = load_WB['Machines_IP'] # 엑셀 시트 명 넣기 (IP)
load_WS2 = load_WB['config_command'] # 엑셀 시트 명 넣기 (config)

# config_command = ["do ter len 0", "do show version", "do show arp",
#                   "do show ip int br", "do show int status", "do show ip route"]

for x in range(1, load_WS1.max_row + 1):
    protocol = load_WS1.cell(row=x, column=1).value
    road_IP = load_WS1.cell(row=x, column=2).value
    road_ID = load_WS1.cell(row=x, column=3).value
    road_PW = load_WS1.cell(row=x, column=4).value
    road_enable = load_WS1.cell(row=x, column=5).value

    try:
        cisco_test = {
            'device_type': protocol,
            'ip': road_IP,
            'username': road_ID,
            'password': road_PW,
            'secret': road_enable,
            }
        terminal = ConnectHandler(**cisco_test)
        terminal.enable()

        result = terminal.find_prompt()
        print('연결된 장비 : ' + road_IP)
        print(result)
        output1 = terminal.send_command("show ip int br | inc up", use_textfsm=True)     
        # print(output)
        l = len(output1)
        print("Total number of interfaces are " + str(l))

        num = range(l)
        for int_num in num:
            interface0 = output1[int_num]
            getintf = itemgetter('intf')
            getstatus = itemgetter('status')
            name = getintf(interface0)
            status = getstatus(interface0)
            print('\n Interface ' + name + ' status is ' + status)

        terminal.disconnect()

    except Exception as error:
        sys.error_text = open(road_IP + "_error.txt", "w")
        print(error, file=sys.error_text)

    print('연결 종료 : ' + road_IP)

