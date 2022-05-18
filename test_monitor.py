from operator import itemgetter
from netmiko import ConnectHandler
from openpyxl import load_workbook, Workbook
import sys
import time
import schedule

load_WB = load_workbook("sample.xlsx") # 엑셀 파일 명 넣기
load_WS1 = load_WB['Machines_IP'] # 엑셀 시트 명 넣기 (IP)
load_WS2 = load_WB['config_command'] # 엑셀 시트 명 넣기 (config)

write_wb = Workbook()
write_ws1 = write_wb.active
write_ws1.title = "monitor_interface"

write_ws1.append(['번호', '시간', 'Interface_name', 'input_packets', 'output_packets'])
def packet_monitor():
    for x in range(1, load_WS1.max_row + 1):
        time_check = time.strftime('%Y-%m-%d %H:%M:%S')

        protocol = load_WS1.cell(row=x, column=1).value
        road_IP = load_WS1.cell(row=x, column=2).value
        road_ID = load_WS1.cell(row=x, column=3).value
        road_PW = load_WS1.cell(row=x, column=4).value
        road_enable = load_WS1.cell(row=x, column=5).value
        try:
            cisco_test = {
                'device_type': protocol,
                'ip': road_IP,
                'username': road_ID,
                'password': road_PW,
                'secret': road_enable,
                }
            terminal = ConnectHandler(**cisco_test)
            terminal.enable()

            result = terminal.find_prompt()
            print('연결된 장비 : ' + road_IP)
            print(result)
            output = terminal.send_command("show int vlan1", use_textfsm=True)

            show_int = output[(0)]
            getname = itemgetter('interface')
            wr_getname = getname(show_int)
            getinput_pk = itemgetter('input_packets')
            wr_input_pk = getinput_pk(show_int)
            getoutput_pk = itemgetter('output_packets')
            wr_output_pk = getoutput_pk(show_int)
            write_ws1.append([x, time_check, wr_getname, wr_input_pk, wr_output_pk])
            print(x, time_check, wr_getname, wr_input_pk, wr_output_pk)
            terminal.disconnect()
            print('연결 종료 : ' + road_IP)
        except Exception as error:
            sys.error_text = open(road_IP + "_error.txt", "w")
            print(error, file=sys.error_text)
        
        write_wb.save('test_monitor.xlsx')
schedule.every(5).seconds.do(packet_monitor) # 5초마다 실행 반복문

while True: # 5초마다 실행 반복문
    schedule.run_pending() # 5초마다 실행 반복문
    time.sleep(1) # 5초마다 실행 반복문
